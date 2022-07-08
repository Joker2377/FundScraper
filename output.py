from FundScraper import *
from openpyxl import *
from openpyxl.styles import Font

if __name__ == '__main__':
    wb = Workbook()
    f = FundScraper()
    f.startFundDetailScraping(byId=True)
    d1 = f.readData()
    print(d1)
    if len(d1) == 1:
        d1 = d1[0]
    sheet = wb.create_sheet(d1['name'])
    word_dict = {'name': '基金名稱',
                 'id': '基金代碼',
                 'eng_name': '基金名稱(英)',
                 'general_agent': '總代理',
                 'company': '基金公司',
                 'create_date': '成立日期',
                 'fund_type': '基金類型',
                 'register_country': '基金註冊地',
                 'goal': '投資目標',
                 'original_scale': '原始可發行規模(百萬)',
                 'scale': '基金規模',
                 'manage_institution': '保管機構',
                 'manage_fee': '基金保管費率',
                 'manage_fee_max': '基金管理費率(最高)',
                 'earning_distribute': '收益分配方式',
                 'buying_handling_fee': '申購手續費',
                 'company_url': '公司簡介網址',
                 'value_list': '淨值表',
                 'earn_list': '績效表現',
                 'configure_list': '資產配置',
                 'industry_proportion': '行業比重',
                 'risk_evaluate': '風險評估',
                 'top10_shareholding': '前十大持股',
                 'risk': '風險評等',
                 'dividend': '配息紀錄'
                 }
    for key, value in d1.items():
        if type(value) is not list:
            tmp = [word_dict[key], value]
            sheet.append(tmp)
            sheet.append([])
        else:
            sheet.append([word_dict[key]])
            for x in value:
                if type(x[1]) is not list:
                    sheet.append(x)
                else:
                    title = x[0]
                    x = x[1]
                    sheet.append(x)
            sheet.append([])
    for col in sheet.iter_cols(min_row=1, max_col=5, max_row=200):
        for cell in col:
            if str(cell.value) in word_dict.values():
                sheet.cell(row=cell.row, column=cell.column).font = Font(color='08ba17',bold=True)  # change title color
            elif '▲' in str(cell.value):
                cell.font = Font(color='ab1717')
            elif '▼' in str(cell.value):
                cell.font = Font(color='3d9414')
    wb.save('test.xlsx')
