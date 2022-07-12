import io  # openpyxl和內置函數open會衝突，所以需指名使用io.open
import json
import sys
import time
from os import system

import selenium
from PyQt5 import QtWidgets
from PyQt5.QtCore import QObject, QThread
from openpyxl import *
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from Fund import *
from UI import Ui_MainWindow

TextEditControl = None


class Worker(QObject):
    def start_process1(self):
        f = FundScraper()
        f.setTarget(3)
        f.startFundListScraping()
        f.startFundDetailScraping()
        f.excelOutput()

    def start_process2(self):
        f = FundScraper()
        f.setTarget(1)
        f.startFundListScraping()
        f.startFundDetailScraping()
        f.excelOutput()

    def start_process3(self):
        f = FundScraper()
        f.setTarget(2)
        f.startFundListScraping()
        f.startFundDetailScraping()
        f.excelOutput()


class MainWindow_controller(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        self.thread = None
        self.worker = None
        global TextEditControl
        TextEditControl = self.ui.textEdit

    def setup_control(self):

        self.ui.comboBox_fundList.addItem('全部')  # add items to first comboBox
        self.ui.comboBox_fundList.addItem('單筆top20')
        self.ui.comboBox_fundList.addItem('定額top20')

        self.ui.pushButton_fundList.clicked.connect(self.onFundListButtonClick)

    def onFundListButtonClick(self):
        s = self.ui.comboBox_fundList.currentText()
        if s == '全部':
            self.start_process(1)
        elif s == '單筆top20':
            self.start_process(2)
        elif s == '定額top20':
            self.start_process(3)

    def message(self, s):
        self.ui.textEdit.append(s)

    def start_process(self, num):
        self.ui.pushButton_fundList.setText('結束')
        if num == 1:
            self.thread = QThread()
            self.worker = Worker()
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.start_process1)
            self.thread.start()
        elif num == 2:
            self.thread = QThread()
            self.worker = Worker()
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.start_process2)
            self.thread.start()
        elif num == 3:
            self.thread = QThread()
            self.worker = Worker()
            self.worker.moveToThread(self.thread)
            self.thread.started.connect(self.worker.start_process3)
            self.thread.start()


class FundScraper:
    def __init__(self):
        self.f1 = None
        self.f2 = None
        self.url1 = 'https://www.fundrich.com.tw/new-theme-fund/root.HOT.hot006'  # 單筆top20
        self.url2 = 'https://www.fundrich.com.tw/new-theme-fund/root.HOT.hot13'  # 定額top20
        self.url3 = 'https://www.fundrich.com.tw/new-theme-fund/'  # 全部
        self.url = None
        self.fundList = []
        self.fundIdList = []
        self.driverInitializeOrNot = False
        self.scrape_type = None

    class FundListScraper:
        def __init__(self):
            self.fundList = []
            self.url = None

        class FundListNetProcess:
            def __init__(self, url):
                self.url = url
                opt = webdriver.EdgeOptions()
                opt.add_argument('log-level=3')
                self.driver = webdriver.Edge(options=opt)
                self.driver.get(self.url)

            def getFundInfo(self):  # 取得清單上的基金資料(代碼、基金名稱)
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'tbody')))
                        element = self.driver.find_element(By.CLASS_NAME, 'tbody')
                        result = element.text
                        str = result.split('立即結帳')
                        funds = []
                        for x in str:
                            s = x.split('\n')
                            newS = []
                            for y in s:
                                if y != '':
                                    newS.append(y)
                            if newS:
                                funds.append([newS[0], newS[1]])
                        if not funds:
                            continue
                        return funds
                    except selenium.common.exceptions.NoSuchElementException:
                        self.escape()  # 關閉廣告
                        TextEditControl.append('---Waiting---')
                        time.sleep(1)
                        self.driver.refresh()
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.escape()
                        self.driver.refresh()
                        continue

            def escape(self):  # 避免廣告視窗干擾運作
                for x in range(3):
                    webdriver.ActionChains(self.driver).move_by_offset(0, 0).click().perform()

            def nextPage(self):  # 下一頁
                times = 0
                while True:
                    try:
                        time.sleep(1)
                        wait = WebDriverWait(self.driver, 5)
                        btn = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "btn-next")))
                        btn.click()
                        element = self.driver.find_element(By.CLASS_NAME, "btn-next")
                        times = 0
                        if element.is_enabled():
                            return True
                        else:
                            return False
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        TextEditControl.append('---Waiting---')
                        times += 1
                        if times > 3:
                            self.escape()
                            TextEditControl.append('Closing ads...')
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        TextEditControl.append('Refreshing the website...')
                        time.sleep(3)
                        continue

            def close(self):  # 關閉瀏覽器視窗
                self.driver.close()

        def setTargetUrl(self, url):
            self.url = url

        def start(self):
            if self.url is None:
                raise ValueError("Url invalid!")
            net = self.FundListNetProcess(self.url)
            funds = net.getFundInfo()
            fundList = []
            for x in funds:
                f = Fund(x[0], x[1])
                fundList.append(f)
            net.nextPage()
            startTime = time.time()
            oldFunds = []
            while funds:
                funds = []
                funds = net.getFundInfo()
                if oldFunds == funds:
                    continue
                oldFunds = funds
                TextEditControl.append(f"---Processing---\nTotal:{len(funds)}\n")
                newFunds = []
                for x in funds:
                    f = Fund(x[0], x[1])
                    newFunds.append(f)
                fundList = fundList + newFunds
                if not net.nextPage():
                    break

            net.close()
            TextEditControl.append(f"Collected: {len(fundList)}")
            self.fundList = fundList
            system('cls')
            return

        def getFundList(self):
            return self.fundList

    class FundDetailScraper:  # earnList not added
        def __init__(self, **kwargs):
            if not kwargs:
                self.id = 'ALI006'  # 參考基金，用以建立driver instance
            else:
                self.id = kwargs['id']
            self.infoUrl = f"https://www.fundrich.com.tw/fund/{self.id}.html?id={self.id}#%E5%9F%BA%E9%87%91%E8%B3%87%E6%96%99"
            self.infoList = []  # 基金資料
            self.earnList = []  # 績效表現
            self.valueList = []  # 近30日淨值表
            self.confList = []  # 資產配置
            self.dataList = []  # 行業比重、風險評估、前十大持股
            self.risk = None  # 風險評等
            self.dividend = None  # 配息紀錄
            self.net = self.FundDetailNetProcess(self.infoUrl)

        class FundDetailNetProcess:
            def __init__(self, url):
                self.url = url
                opt = webdriver.EdgeOptions()
                opt.add_argument('log-level=3')
                opt.headless = True
                self.driver = webdriver.Edge(options=opt)
                self.driver.get(self.url)
                TextEditControl.append("---Waiting---")
                time.sleep(3)
                system('cls')

            def setTarget(self, url):
                self.url = url
                self.driver.get(url)

            def getFundDetail(self):  # 基金資料
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'col-sm-8')))
                        elements = self.driver.find_elements(By.CLASS_NAME, 'col-sm-8')

                        infoList = []
                        for x in elements:
                            str = x.text.split('\n')
                            infoList += str
                        infoList = infoList[:-7:]

                        elements = self.driver.find_elements(By.CSS_SELECTOR, 'td > a')  # 公司網址
                        href = [elem.get_attribute('href') for elem in elements]

                        infoList += href
                        return infoList
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        self.escape()
                        continue
                    except selenium.common.exceptions.NoSuchElementException:
                        TextEditControl.append('---Waiting---')
                        time.sleep(1)
                        self.driver.refresh()
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        continue

            def getFundValueList(self):  # 淨值表
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        btn = wait.until(EC.presence_of_element_located((By.XPATH, "//span[text()='淨值走勢']")))
                        btn.click()
                        dateElements = self.driver.find_elements(By.XPATH,
                                                                 "//td[@style='padding: 3px; height: 48px; text-align: center; "
                                                                 "font-size: 14px; overflow: hidden; white-space: inherit; "
                                                                 "text-overflow: inherit; background-color: rgb(240, 240, "
                                                                 "237); word-break: break-all; font-weight: 500; line-height: "
                                                                 "2.95; letter-spacing: 0.7px; color: rgb(119, 119, 119);']")
                        valueElements = self.driver.find_elements(By.XPATH,
                                                                  "//td[@style='padding: 3px; height: 48px; text-align: "
                                                                  "center; font-size: 14px; overflow: hidden; white-space: "
                                                                  "inherit; text-overflow: inherit; background-color: "
                                                                  "inherit; word-break: break-all; font-weight: 500; "
                                                                  "line-height: 2.95; letter-spacing: 0.7px; color: rgb("
                                                                  "119, 119, 119);']")
                        dateList = []
                        valueList = []
                        for x in dateElements:
                            dateList.append(str(x.text))
                        for x in valueElements:
                            valueList.append(str(x.text))
                        c = list(zip(dateList, valueList))
                        if not c:
                            self.driver.refresh()
                            continue
                        return c
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        self.escape()
                        continue
                    except selenium.common.exceptions.NoSuchElementException:
                        TextEditControl.append('---Waiting---')
                        time.sleep(1)
                        self.driver.refresh()
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        continue

            def getEarn(self):  # 績效表現
                wait = WebDriverWait(self.driver, 5)
                btn = wait.until(EC.presence_of_element_located((By.XPATH, "//span[text()='績效表現']")))
                btn.click()
                element = self.driver.find_element(By.XPATH,
                                                   "//div[@style='width: 100%; padding-bottom: 45px; display: "
                                                   "inline-block;']//div[@class='row']")
                ls = element.text.split('\n')
                re = []
                for x in ls:
                    c = x.split()
                    if tuple(c) not in re:
                        if len(c) == 3:
                            c = [c[0], c[1] + c[2]]
                        c = tuple(c)
                        re.append(c)
                return re

            def getFundConfigure(self):  # 資產配置
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        btn = wait.until(EC.presence_of_element_located((By.XPATH, "//span[text()='資產配置']")))
                        btn.click()
                        elements1 = self.driver.find_elements(By.XPATH, "//*[@text-anchor='end']")
                        elements2 = self.driver.find_elements(By.XPATH,
                                                              "//*[@style='background-color: rgb(255, 255, 255); padding: "
                                                              "0px 24px; width: 100%; border-collapse: collapse; "
                                                              "border-spacing: 0px; table-layout: fixed; font-family: "
                                                              "微軟正黑體, \"Microsoft JhengHei\", SimHei, 新細明體, Arial, "
                                                              "Verdana, Helvetica, sans-serif;']")

                        data = [elements1, elements2]  # 資產配置(股票現金占比)/資產配置頁的其他東西
                        confList = []
                        dataList = []
                        for x in data[0]:
                            tmp = x.text.split()
                            c = tuple([tmp[0], tmp[1]])
                            confList.append(c)
                        for x in data[1]:
                            if x.text != '':
                                s = x.text.split('\n')
                                dataList.append(s)
                        newDataList = []
                        lst = []
                        for x in dataList[0]:
                            if dataList[0].index(x) > 0:
                                tmp = x.split()
                                value = tmp[-1]
                                tmp = tmp[:-1:]
                                str1 = ""
                                for y in tmp:
                                    str1 += y + " "
                                c = tuple([str1, value])
                                lst.append(c)
                            else:
                                c = tuple(x.split())
                                lst.append(c)
                        newDataList.append(lst)
                        lst = []
                        for x in dataList[1]:
                            if dataList[1].index(x) > 0:
                                tmp = x.split()
                                unit = tmp[0]
                                value = tmp[1::]
                                c = tuple([unit, value])
                                lst.append(c)
                            else:
                                c = tuple(x.split())
                                lst.append(c)
                        newDataList.append(lst)
                        lst = []
                        for x in dataList[2]:
                            if dataList[2].index(x) > 0:
                                tmp = x.split()
                                value = tmp[-1]
                                tmp = tmp[:-1:]
                                str1 = ""
                                for y in tmp:
                                    str1 += y + " "
                                c = tuple([str1, value])
                                lst.append(c)
                            else:
                                c = tuple(x.split())
                                lst.append(c)
                        newDataList.append(lst)
                        return [confList, newDataList]
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        self.escape()
                        continue
                    except selenium.common.exceptions.NoSuchElementException:
                        TextEditControl.append('---Waiting---')
                        time.sleep(1)
                        self.driver.refresh()
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        continue

            def getFundRisk(self):  # 風險評等
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        btn = wait.until(EC.presence_of_element_located((By.XPATH, "//span[text()='風險評等']")))
                        btn.click()
                        element = self.driver.find_element(By.XPATH,
                                                           "//*[@style='margin: 5px 0px; text-align: center; font-size: "
                                                           "18px; font-weight: 500; line-height: 1.67;']")
                        return element.text
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        self.escape()
                        continue
                    except selenium.common.exceptions.NoSuchElementException:
                        TextEditControl.append('---Waiting---')
                        time.sleep(1)
                        self.driver.refresh()
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        continue

            def getFundDividend(self):  # 配息紀錄
                while True:
                    try:
                        wait = WebDriverWait(self.driver, 5)
                        btn = wait.until(EC.presence_of_element_located((By.XPATH, "//span[text()='配息紀錄']")))
                        btn.click()
                        element = self.driver.find_element(By.XPATH, "//*[contains(text(), '無配息資料')]")
                        if element.get_attribute('style') == 'display: none;':
                            elements = self.driver.find_elements(By.XPATH,
                                                                 "//*[@style='background-color: rgb(255, 255, 255); padding: 0px "
                                                                 "24px; width: 100%; border-collapse: collapse; border-spacing: 0px; "
                                                                 "table-layout: fixed; font-family: 微軟正黑體, \"Microsoft JhengHei\", "
                                                                 "SimHei, 新細明體, Arial, Verdana, Helvetica, sans-serif;']")
                            b = None
                            for elem in elements:
                                a = elem.text.split('\n')
                                b = []
                                for item in a:
                                    c = item.split()
                                    if c:
                                        c = tuple(c)
                                        b.append(c)

                            return b
                        else:
                            return '無配息資料'
                    except selenium.common.exceptions.ElementClickInterceptedException:
                        self.escape()
                        continue
                    except selenium.common.exceptions.NoSuchElementException:
                        TextEditControl.append('---Waiting---')
                        time.sleep(3)
                        continue
                    except selenium.common.exceptions.TimeoutException:
                        self.driver.refresh()
                        continue

            def escape(self):
                for x in range(2):
                    webdriver.ActionChains(self.driver).move_by_offset(0, 0).click().perform()

            def close(self):
                self.driver.close()

        def setTargetFund(self, id):
            self.id = id
            self.infoUrl = f"https://www.fundrich.com.tw/fund/{self.id}.html?id={self.id}#%E5%9F%BA%E9%87%91%E8%B3%87%E6%96%99"

        def start(self):
            if self.id is None:
                raise ValueError("Invalid ID")
            self.net.setTarget(self.infoUrl)
            self.infoList = self.net.getFundDetail()
            self.earnList = self.net.getEarn()
            self.valueList = self.net.getFundValueList()
            self.confList = self.net.getFundConfigure()[0]
            self.dataList = self.net.getFundConfigure()[1]
            self.risk = self.net.getFundRisk()
            self.dividend = self.net.getFundDividend()
            TextEditControl.append(f"---Data received---{self.id}{self.infoList[0]}")

        def getInfoList(self):
            return self.infoList

        def getEarnList(self):
            return self.earnList

        def getValueList(self):
            return self.valueList

        def getConflist(self):
            return self.confList

        def getDataList(self):
            return self.dataList

        def getRisk(self):
            return self.risk

        def getDividend(self):
            return self.dividend

        def close(self):
            self.net.close()

    def clear(self):
        self.__init__()

    def initializeDriver(self):
        self.f1 = self.FundListScraper()
        self.f2 = self.FundDetailScraper()
        self.driverInitializeOrNot = True

    def setTarget(self, num):
        TextEditControl.append('Target set: ')
        if num == 1:
            TextEditControl.append('單筆top20')
            self.url = self.url1
            self.scrape_type = '單筆top20'
        elif num == 2:
            TextEditControl.append('定額top20')
            self.url = self.url2
            self.scrape_type = '定額top20'
        elif num == 3:
            TextEditControl.append('全部')
            self.url = self.url3
            self.scrape_type = '全部'

    def startFundListScraping(self):
        if not self.driverInitializeOrNot:
            self.initializeDriver()
        self.f1.setTargetUrl(self.url)
        self.f1.start()
        self.fundList = self.f1.getFundList()
        fundIdList = []
        for x in self.fundList:
            fundIdList.append(x.getId())
        self.fundIdList = fundIdList
        self.writeFundIdList(fundIdList)

    def startFundDetailScraping(self, **kwargs):
        if not self.driverInitializeOrNot:
            self.initializeDriver()

        if not self.fundList and not self.fundIdList:
            self.fundIdList = self.readFundIdList()

        if len(kwargs) < 1:
            for x in self.fundList:
                id = x.getId()
                self.f2.setTargetFund(id)
                self.f2.start()

                infoList = self.f2.getInfoList()
                if x.name is None:
                    x.setName(infoList[0])
                x.setValueList(self.f2.getValueList())
                x.setEarnList(self.f2.getEarnList())
                x.setConfList(self.f2.getConflist())
                x.setDataList(self.f2.getDataList())
                x.setRisk(self.f2.getRisk())
                x.setDividend(self.f2.getDividend())
                for info in infoList:
                    x.addInfo(info)
            self.writeData()

        elif kwargs['byId']:
            fundList = []
            for x in self.fundIdList:
                fundList.append(Fund(x))
            self.fundList = fundList
            self.startFundDetailScraping()

    def getFundList(self):
        return self.fundList

    def writeData(self):
        jsonStr = self.toJson(self.fundList)
        self.writeJsonFile(jsonStr)

    def readData(self):
        d1 = self.toDict(self.readJsonFile())  # returning list of dicts
        return d1

    def writeFundIdList(self, fundIdList):
        with io.open('fund_id_list.txt', 'w', newline='') as file:
            for x in fundIdList:
                file.write(x)
                file.write('\n')

    def readFundIdList(self):
        try:
            with io.open('fund_id_list.txt', 'r') as file:
                fundIdList = []
                for x in file.readlines():
                    fundIdList.append(x)
                fundIdList = [x[:-1:] for x in fundIdList]
                return fundIdList
        except FileNotFoundError:
            with io.open('fund_id_list.txt', 'x', newline='') as file:
                pass

    def toJson(self, fundList):
        jsonFundList = []
        for fund in fundList:
            d1 = {'name': fund.name,
                  'id': fund.id,  # 基金代碼
                  'eng_name': fund.infoList[1],  # 英文名稱
                  'general_agent': fund.infoList[2],  # 總代理
                  'company': fund.infoList[3],  # 基金公司
                  'create_date': fund.infoList[4],  # 成立日期
                  'fund_type': fund.infoList[5],  # 基金類型
                  'register_country': fund.infoList[6],  # 基金註冊地
                  'goal': fund.infoList[7],  # 投資目標
                  'original_scale': fund.infoList[8],  # 原始可發行規模(百萬)
                  'scale': fund.infoList[9],  # 基金規模
                  'manage_institution': fund.infoList[10],  # 保管機構
                  'manage_fee': fund.infoList[11],  # 基金保管費率
                  'manage_fee_max': fund.infoList[12],  # 基金管理費率(最高)
                  'earning_distribute': fund.infoList[13],  # 收益分配方式
                  'buying_handling_fee': fund.infoList[14],  # 申購手續費
                  'company_url': fund.infoList[15],  # 公司簡介網址
                  'value_list': fund.valueList,  # 淨值表
                  'earn_list': fund.earnList,  # 績效表現
                  'configure_list': fund.confList,  # 資產配置
                  'industry_proportion': fund.dataList[0],  # 行業比重
                  'risk_evaluate': fund.dataList[1],  # 風險評估
                  'top10_shareholding': fund.dataList[2],  # 前十大持股
                  'risk': fund.risk,  # 風險評等(RR)
                  'dividend': fund.dividend  # 配息紀錄
                  }
            jsonFundList.append(d1)
        jsonStr = json.dumps(jsonFundList, ensure_ascii=False)
        return jsonStr

    def toDict(self, jsonStr):
        j = json.loads(jsonStr)
        return j

    def writeJsonFile(self, jsonStr):
        with io.open('fund_list.json', 'w', encoding='UTF-8') as jsonFile:
            jsonFile.write(str(jsonStr))

    def readJsonFile(self):
        try:
            with io.open('fund_list.json', 'r', encoding='UTF-8') as jsonFile:
                jsonStr = jsonFile.read()
                return jsonStr
        except FileNotFoundError:
            with io.open('fund_list.json', 'x', encoding='UTF-8', newline='') as file:
                pass

    def excelOutput(self):
        d1 = self.readData()
        word_dict = {'name': '基金名稱',  # 將英文資料名稱對應成中文
                     'id': '基金代碼',
                     'eng_name': '基金名稱(英)',
                     'general_agent': '總代理',
                     'company': '基金公司',
                     'create_date': '成立日期',
                     'fund_type': '基金類型',
                     'register_country': '基金註冊地',
                     'goal': '投資目標',
                     'original_scale': '原始可發行規模(百萬)',
                     'scale': '基金規模',
                     'manage_institution': '保管機構',
                     'manage_fee': '基金保管費率',
                     'manage_fee_max': '基金管理費率(最高)',
                     'earning_distribute': '收益分配方式',
                     'buying_handling_fee': '申購手續費',
                     'company_url': '公司簡介網址',
                     'value_list': '淨值表',
                     'earn_list': '績效表現',
                     'configure_list': '資產配置',
                     'industry_proportion': '行業比重',
                     'risk_evaluate': '風險評估',
                     'top10_shareholding': '前十大持股',
                     'risk': '風險評等',
                     'dividend': '配息紀錄'
                     }

        wb = Workbook()
        for x in d1:
            sheetName = x['id'] + ' ' + x['name']
            if len(sheetName) > 31:
                sheetName = sheetName[:31:]
            for sheet in wb:  # 避免重複
                ws = sheet
                if ws.title == sheetName:
                    wb.remove(wb[sheetName])
            ws = wb.create_sheet(sheetName)
            for key, value in x.items():
                if type(value) is not list:  # 判斷資料是否為list
                    tmp = [word_dict[key], value]
                    ws.append(tmp)
                    ws.append([])  # 空行
                else:
                    ws.append([word_dict[key]])
                    for y in value:
                        if type(y[1]) is not list:
                            ws.append(y)
                        else:
                            title = y[0]
                            y = y[1]
                            ws.append(y)
                    ws.append([])
            for col in ws.iter_cols(min_row=1, max_col=5, max_row=200):  # adding styles
                for cell in col:
                    if str(cell.value) in word_dict.values():
                        ws.cell(row=cell.row, column=cell.column).font = Font(color='08ba17',
                                                                              bold=True)  # change title color
                    elif '▲' in str(cell.value):
                        cell.font = Font(color='ab1717')
                    elif '▼' in str(cell.value):
                        cell.font = Font(color='3d9414')
        re = time.localtime(time.time())

        wb.save(f'{re.tm_wday}{re.tm_hour}{re.tm_min}{re.tm_sec}_{self.scrape_type}_result.xlsx')
        TextEditControl.append('Process finished.')

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow_controller()
    window.show()
    sys.exit(app.exec_())
