import json
from FundDetailScraper import *
from FundListScraper import *
from openpyxl import *
from openpyxl.styles import Font
import io  # openpyxl和內置函數open會衝突，所以需指名使用io.open
import time


class FundScraper:
    def __init__(self):
        self.f1 = None
        self.f2 = None
        self.url1 = 'https://www.fundrich.com.tw/new-theme-fund/root.HOT.hot006'  # 單筆top20
        self.url2 = 'https://www.fundrich.com.tw/new-theme-fund/root.HOT.hot13'  # 定額top20
        self.url3 = 'https://www.fundrich.com.tw/new-theme-fund/'  # 全部
        self.url = None
        self.fundList = []
        self.fundIdList = []
        self.driverInitializeOrNot = False
        self.scrape_type = None

    def clear(self):
        self.__init__()

    def initializeDriver(self):
        self.f1 = FundListScraper()
        self.f2 = FundDetailScraper()
        self.driverInitializeOrNot = True

    def setTarget(self, num):
        print('Target set: ', end='')
        if num == 1:
            print('單筆top20')
            self.url = self.url1
            self.scrape_type = '單筆top20'
        elif num == 2:
            print('定額top20')
            self.url = self.url2
            self.scrape_type = '定額top20'
        elif num == 3:
            print('全部')
            self.url = self.url3
            self.scrape_type = '全部'
        else:
            print('nothing')

    def startFundListScraping(self):
        if not self.driverInitializeOrNot:
            self.initializeDriver()
        self.f1.setTargetUrl(self.url)
        self.f1.start()
        self.fundList = self.f1.getFundList()
        fundIdList = []
        for x in self.fundList:
            fundIdList.append(x.getId())
        self.fundIdList = fundIdList
        self.writeFundIdList(fundIdList)

    def startFundDetailScraping(self, **kwargs):
        if not self.driverInitializeOrNot:
            self.initializeDriver()

        if not self.fundList and not self.fundIdList:
            print('Get a fundList or fundIdList before starting \"fund detail scraping\".')

        if len(kwargs) < 1:
            if not self.fundList:
                print('Get a fundList before starting \"fund detail scraping by fundList\".')
            for x in self.fundList:
                id = x.getId()
                self.f2.setTargetFund(id)
                self.f2.start()

                infoList = self.f2.getInfoList()
                if x.name is None:
                    x.setName(infoList[0])
                x.setValueList(self.f2.getValueList())
                x.setEarnList(self.f2.getEarnList())
                x.setConfList(self.f2.getConflist())
                x.setDataList(self.f2.getDataList())
                x.setRisk(self.f2.getRisk())
                x.setDividend(self.f2.getDividend())
                for info in infoList:
                    x.addInfo(info)
            self.writeData()

        elif kwargs['byId']:
            fundList = []
            for x in self.fundIdList:
                fundList.append(Fund(x))
            self.fundList = fundList
            self.startFundDetailScraping()

    def getFundList(self):
        return self.fundList

    def writeData(self):
        jsonStr = self.toJson(self.fundList)
        self.writeJsonFile(jsonStr)

    def readData(self):
        d1 = self.toDict(self.readJsonFile())  # returning list of dicts
        return d1

    def writeFundIdList(self, fundIdList):
        with io.open('fund_id_list.txt', 'w', newline='') as file:
            for x in fundIdList:
                file.write(x)
                file.write('\n')

    def readFundIdList(self):
        try:
            with io.open('fund_id_list.txt', 'r') as file:
                fundIdList = []
                for x in file.readlines():
                    fundIdList.append(x)
                fundIdList = [x[:-1:] for x in fundIdList]
                return fundIdList
        except FileNotFoundError:
            with io.open('fund_id_list.txt', 'x', newline='') as file:
                pass

    def toJson(self, fundList):
        jsonFundList = []
        for fund in fundList:
            d1 = {'name': fund.name,
                  'id': fund.id,  # 基金代碼
                  'eng_name': fund.infoList[1],  # 英文名稱
                  'general_agent': fund.infoList[2],  # 總代理
                  'company': fund.infoList[3],  # 基金公司
                  'create_date': fund.infoList[4],  # 成立日期
                  'fund_type': fund.infoList[5],  # 基金類型
                  'register_country': fund.infoList[6],  # 基金註冊地
                  'goal': fund.infoList[7],  # 投資目標
                  'original_scale': fund.infoList[8],  # 原始可發行規模(百萬)
                  'scale': fund.infoList[9],  # 基金規模
                  'manage_institution': fund.infoList[10],  # 保管機構
                  'manage_fee': fund.infoList[11],  # 基金保管費率
                  'manage_fee_max': fund.infoList[12],  # 基金管理費率(最高)
                  'earning_distribute': fund.infoList[13],  # 收益分配方式
                  'buying_handling_fee': fund.infoList[14],  # 申購手續費
                  'company_url': fund.infoList[15],  # 公司簡介網址
                  'value_list': fund.valueList,  # 淨值表
                  'earn_list': fund.earnList,  # 績效表現
                  'configure_list': fund.confList,  # 資產配置
                  'industry_proportion': fund.dataList[0],  # 行業比重
                  'risk_evaluate': fund.dataList[1],  # 風險評估
                  'top10_shareholding': fund.dataList[2],  # 前十大持股
                  'risk': fund.risk,  # 風險評等(RR)
                  'dividend': fund.dividend  # 配息紀錄
                  }
            jsonFundList.append(d1)
        jsonStr = json.dumps(jsonFundList, ensure_ascii=False)
        return jsonStr

    def toDict(self, jsonStr):
        j = json.loads(jsonStr)
        return j

    def writeJsonFile(self, jsonStr):
        with io.open('fund_list.json', 'w', encoding='UTF-8') as jsonFile:
            jsonFile.write(str(jsonStr))

    def readJsonFile(self):
        try:
            with io.open('fund_list.json', 'r', encoding='UTF-8') as jsonFile:
                jsonStr = jsonFile.read()
                return jsonStr
        except FileNotFoundError:
            with io.open('fund_list.json', 'x', encoding='UTF-8', newline='') as file:
                pass

    def excelOutput(self):
        d1 = self.readData()
        word_dict = {'name': '基金名稱',  # 將英文資料名稱對應成中文
                     'id': '基金代碼',
                     'eng_name': '基金名稱(英)',
                     'general_agent': '總代理',
                     'company': '基金公司',
                     'create_date': '成立日期',
                     'fund_type': '基金類型',
                     'register_country': '基金註冊地',
                     'goal': '投資目標',
                     'original_scale': '原始可發行規模(百萬)',
                     'scale': '基金規模',
                     'manage_institution': '保管機構',
                     'manage_fee': '基金保管費率',
                     'manage_fee_max': '基金管理費率(最高)',
                     'earning_distribute': '收益分配方式',
                     'buying_handling_fee': '申購手續費',
                     'company_url': '公司簡介網址',
                     'value_list': '淨值表',
                     'earn_list': '績效表現',
                     'configure_list': '資產配置',
                     'industry_proportion': '行業比重',
                     'risk_evaluate': '風險評估',
                     'top10_shareholding': '前十大持股',
                     'risk': '風險評等',
                     'dividend': '配息紀錄'
                     }

        wb = Workbook()
        for x in d1:
            sheetName = x['id'] + ' ' + x['name']
            for sheet in wb:  # 避免重複
                ws = sheet
                if ws.title == sheetName:
                    wb.remove(wb[sheetName])
            ws = wb.create_sheet(sheetName)
            for key, value in x.items():
                if type(value) is not list:  # 判斷資料是否為list
                    tmp = [word_dict[key], value]
                    ws.append(tmp)
                    ws.append([])  # 空行
                else:
                    ws.append([word_dict[key]])
                    for y in value:
                        if type(y[1]) is not list:
                            ws.append(y)
                        else:
                            title = y[0]
                            y = y[1]
                            ws.append(y)
                    ws.append([])
            for col in ws.iter_cols(min_row=1, max_col=5, max_row=200):  # adding styles
                for cell in col:
                    if str(cell.value) in word_dict.values():
                        ws.cell(row=cell.row, column=cell.column).font = Font(color='08ba17',
                                                                              bold=True)  # change title color
                    elif '▲' in str(cell.value):
                        cell.font = Font(color='ab1717')
                    elif '▼' in str(cell.value):
                        cell.font = Font(color='3d9414')
        re = time.localtime(time.time())

        wb.save(f'{re.tm_mon}{re.tm_mday}{re.tm_hour}{re.tm_sec}{re.tm_wday}_{self.scrape_type}_scrape_result.xlsx')
